import pandas as pd
from opt_deployment_ammonia import main as opt_ammonia
from opt_deployment_refining import main as opt_refining
from opt_deployment_steel import main as opt_steel
from pp_opt_sectors import main as pp_opt
from process_heat_pp import main as process_heat_launch_pp
from utils import WACC
from openpyxl import load_workbook

def launch_opt(noak_tag, itc):
    # check if results already exist
    wb = load_workbook(f'./results/raw_results_SMR_{noak_tag}_ITC_{itc}.xlsx', read_only=True)
    if 'refining' not in wb.sheetnames:
        print("Refining optimization")
        opt_refining(SMR_tag=noak_tag, ITC=itc)
    elif "ammonia" not in wb.sheetnames:
        print("Ammonia optimization")
        opt_ammonia(SMR_tag=noak_tag, ITC=itc)
    elif "steel" not in wb.sheetnames:
        print("Steel optimization")   
        opt_steel(SMR_tag=noak_tag, ITC=itc)
    else:
        print("Results already exist, proceeding to post-processing \n")

def pp_noak_results(noak_tag, itc):
    pp_opt(OAK=noak_tag, ITC=itc, wacc=WACC)


def main():
    print('NOAK deployment phase \n')
    for noak_tag, inc in {'NOAK_with_inc':[True,0.3], 'NOAK_wo_inc':[False,0]}.items():
        print('Launch optimization for ammonia, refining and steel sectors\n')
        launch_opt(noak_tag=noak_tag, itc=inc[1])
        print('Post process results from optimization of ammonia, refining and steel sectors\n')
        pp_noak_results(noak_tag=noak_tag, itc=inc[1])
        print('Launch analysis and post processing of SMR deployment for process heat\n')
        process_heat_launch_pp(OAK=noak_tag, with_PTC=inc[0], cogen=True, ITC=inc[1])


if __name__ == '__main__':
    main()
